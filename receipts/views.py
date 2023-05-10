from django.shortcuts import render
from django.db.models import F, Q, CharField, Value, Sum
import operator
from functools import reduce
from datetime import datetime, date
from django.contrib.postgres.aggregates import ArrayAgg
from django.db.models.functions import Concat
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponseRedirect, JsonResponse
from babel.dates import format_date
import calendar
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import FileResponse
from django.http import HttpResponse
from tempfile import NamedTemporaryFile
from django.utils.encoding import smart_str
from django.template.loader import render_to_string
from io import BytesIO, StringIO
from django.template.loader import get_template




from openpyxl import load_workbook
from openpyxl.styles import Side, Border, Font, Alignment, NamedStyle
from xhtml2pdf import pisa


from decimal import Decimal


from django.views.generic.base import TemplateView
from django.views.generic.detail import DetailView
from django.views.generic.edit import FormView
from users.models import User
from appartments.models import House, Section, Appartment, PersonalAccount
from utility_services.models import Tariff, TariffCell, CounterReadings
from .models import Receipt, ReceiptTemplate, ReceiptCell
from .forms import AddReceiptForm, UtilityReceiptForm, ReceiptCellFormset, ReceiptTemplateListForm, ReceiptTeplateEditeFormSet, ReceiptTeplateEditeForm



class CRMReportView(TemplateView):
    
    template_name = "receipts/crm_report.html"



class ReceiptListView(TemplateView):
    template_name = 'receipts/receipt_list.html'

    def get(self, request, *args, **kwargs):


        # SELECT2 logic
        if self.request.is_ajax() and self.request.GET.get('issue_marker') == 'owners':
            if self.request.GET.get('search'):
                search_data = self.request.GET.get('search')
                owners_data = list(User.objects.filter(full_name__icontains=search_data).values('id', 'full_name'))
                for owner_dict in owners_data: owner_dict['text'] = owner_dict.pop('full_name')
                data = {'results': owners_data}
                return JsonResponse(data)
                

        # users search using Select2
        if self.request.is_ajax() and self.request.GET.get('issue_marker') == 'all_owners':
            owners_data = list(User.objects.all().values('id', 'full_name'))
            for owner_dict in owners_data: owner_dict['text'] = owner_dict.pop('full_name')
            data = {'results': owners_data}
            return JsonResponse(data)



       # datatables serverside logic
        if self.request.is_ajax() and self.request.method == 'GET' and request.GET.get('draw'):
            receipt_data_get_request = request.GET

            #search logic 
            Q_list = []

            # number filtering
            if request.GET.get('columns[1][search][value]'):
                Q_list.append(Q(number__icontains=request.GET.get('columns[1][search][value]')))

            # status filtering
            if request.GET.get('columns[2][search][value]'):
                print(request.GET.get('columns[2][search][value]'))
                if request.GET.get('columns[2][search][value]') != 'all_status':
                    Q_list.append(Q(status=request.GET.get('columns[2][search][value]')))

            # date range search
            if request.GET.get('columns[3][search][value]'):

                date_list = request.GET.get('columns[3][search][value]').split('-')

                start_date = date_list[0].strip().split('.')
                finish_date = date_list[1].strip().split('.')
                start_date.reverse()
                finish_date.reverse()

                start_date_string = ''
                finish_date_string = ''
                
                start_date_string = '-'.join(str(elem) for elem in start_date)
                finish_date_string = '-'.join(str(elem) for elem in finish_date)

                formated_date_start = datetime.strptime(start_date_string, '%Y-%m-%d')
                formated_date_finish = datetime.strptime(finish_date_string, '%Y-%m-%d')
                Q_list.append(Q(to_date__gte=formated_date_start))
                Q_list.append(Q(to_date__lte=formated_date_finish))

            # date with month and year
            if request.GET.get('columns[4][search][value]'):
                date_list = request.GET.get('columns[4][search][value]').split('.')
                month = int(date_list[0])
                year = int(date_list[1])
                first_last_day = calendar.monthrange(year, month)
                first_day = date(year, month, 1)
                last_day = date(year, month, first_last_day[1])
                Q_list.append(Q(date_with_month_year__gte=first_day))
                Q_list.append(Q(date_with_month_year__lte=last_day))

            # full address filter
            if request.GET.get('columns[5][search][value]'):
                search_address_list_parameter = list((request.GET.get('columns[5][search][value]').strip()).split(" "))
                Q_list.append(reduce(operator.and_, (Q(appartments_address__icontains=part_addr) for part_addr in search_address_list_parameter)))

            # full name of owner
            if request.GET.get('columns[6][search][value]'):
                print('----------------------------')
                print(request.GET.get('columns[6][search][value]'))
                print('----------------------------')
                if request.GET.get('columns[6][search][value]'):
                    print(request.GET.get('columns[6][search][value]'))
                    user = User.objects.get(id=request.GET.get('columns[6][search][value]'))
                    Q_list.append(Q(appartment__owner_user=user))


            # payment status filter
            if request.GET.get('columns[7][search][value]'):
                if request.GET.get('columns[7][search][value]') != 'all_payment_status':
                    Q_list.append(Q(payment_was_made=request.GET.get('columns[7][search][value]')))



            draw = int(receipt_data_get_request.get("draw"))
            start = int(receipt_data_get_request.get("start"))
            length = int(receipt_data_get_request.get("length"))

            raw_data = Receipt.objects.annotate(appartments_address = ArrayAgg(Concat(F('appartment__number'),
                                                                                    Value(', '),
                                                                                    F('appartment__house__title'),
                                                                                    output_field=CharField())
                                                                                    ,distinct=True))\
                                    .annotate(date_with_month_year = F('to_date'))\
                                    .filter(*Q_list)\
                                    .order_by('-to_date')\
                                    .values('number', 'status', 'to_date',\
                                            'date_with_month_year', 'appartments_address',\
                                            'appartment__owner_user__full_name',\
                                            'payment_was_made', 'total_sum', "id")

            data = list(raw_data)
            verbose_status_dict = Receipt.get_verbose_status_dict()

            for receipt in data:  
                verbose_status = ""
                try: 
                    verbose_status = verbose_status_dict[receipt['status']]
                    receipt['status'] = verbose_status
                except:
                    receipt['status'] = ''


                simple_date = receipt['to_date']
                new_simple_date = format_date(simple_date, 'dd.MM.yyyy', locale='ru')
                receipt['to_date'] = new_simple_date

                date_per_month = receipt['date_with_month_year']
                new_date = format_date(date_per_month, 'LLLL Y', locale='ru')
                receipt['date_with_month_year'] = new_date

                payment_status = receipt['payment_was_made']
                if payment_status == True:
                    receipt['payment_was_made'] = 'Проведена'
                else:
                    receipt['payment_was_made'] = 'Не проведена'



            # paginator here
            paginator = Paginator(data, length)
            page_number = start / length + 1
            try:
                obj = paginator.page(page_number).object_list
            except PageNotAnInteger:
                obj = paginator(1).object_list
            except EmptyPage:
                obj = paginator.page(1).object_list

            total = len(data)
            records_filter = total

            response = {
                'data': obj,
                'draw': draw,
                'recordsTotal:': total,
                'recordsFiltered': records_filter,
            }
            return JsonResponse(response, safe=False)
        
        else:
            context = self.get_context_data(**kwargs)
            return self.render_to_response(context)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context



# class AddCounterReadingsView(CreateView):
class AddReceiptView(TemplateView):

    template_name = 'receipts/receipt_create.html'    
    form_class = AddReceiptForm
    model = Receipt
    success_url = reverse_lazy('receipts:receipt_list')


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['main_form'] = AddReceiptForm(prefix='main_form')
        context['utility_form'] = UtilityReceiptForm(prefix='utility_form')
        context['receipt_cell_formset'] = ReceiptCellFormset(prefix='receipt_cell_formset')
        return context


    def get(self, request, *args, **kwargs):
        if self.request.is_ajax() and self.request.method == 'GET' and request.GET.get('draw'):
            receipt_data_get_request = request.GET
            Q_list = []
            if request.GET.get('search[value]'):
                if request.GET.get('search[value]') != "empty_appartment":
                    Q_list.append(Q(appartment__id=request.GET.get('search[value]')))

            draw = int(receipt_data_get_request.get("draw"))
            start = int(receipt_data_get_request.get("start"))
            length = int(receipt_data_get_request.get("length"))

            raw_data = CounterReadings.objects.filter(*Q_list)\
                                    .annotate(date_with_month_year = F('date'))\
                                    .values('number', 'status', 'date',\
                                            'date_with_month_year', 'appartment__house__title',\
                                            'appartment__sections__title', 'appartment__number',\
                                            'utility_service__title', 'readings',\
                                            'utility_service__unit_of_measure__title')
            

            data = list(raw_data)
            print(data)
            verbose_status_dict = CounterReadings.get_verbose_status_dict()

            for counter_receipt in data:  
                verbose_status = ""
                try: 
                    verbose_status = verbose_status_dict[counter_receipt['status']]
                    counter_receipt['status'] = verbose_status
                except:
                    counter_receipt['status'] = ''


                simple_date = counter_receipt['date']
                new_simple_date = format_date(simple_date, 'dd.MM.yyyy', locale='ru')
                counter_receipt['date'] = new_simple_date

                date_per_month = counter_receipt['date_with_month_year']
                date_per_month = format_date(date_per_month, 'LLLL Y', locale='ru')
                counter_receipt['date_with_month_year'] = date_per_month

            
            paginator = Paginator(data, length)
            page_number = start / length + 1
            try:
                obj = paginator.page(page_number).object_list
            except PageNotAnInteger:
                obj = paginator(1).object_list
            except EmptyPage:
                obj = paginator.page(1).object_list

            total = len(data)
            records_filter = total

            response = {
                'data': obj,
                'draw': draw,
                'recordsTotal:': total,
                'recordsFiltered': records_filter,
            }
            return JsonResponse(response, safe=False)
        

        if self.request.is_ajax() and self.request.method == 'GET' and self.request.GET.get('ajax_indicator') == 'get_certain_house':
            house_id = self.request.GET['current_house_number']
            house = House.objects.get(id=house_id)
            sections = list(Section.objects.only('id', 'title').filter(house=house).values('id', 'title'))
            appartments = list(Appartment.objects.only('id', 'number').filter(house=house).values('id', 'number'))
            response = {'sections': sections,
                        'appartments':appartments}
            return JsonResponse(response, safe=False)


        if self.request.is_ajax() and self.request.method == 'GET' and self.request.GET.get('ajax_indicator') == 'get_appartments_per_sections':
            sections_id = self.request.GET.get('current_sections_number')
            choosen_sections = Section.objects.get(id=sections_id)
            appartments = list(Appartment.objects.only('id', 'number').filter(sections=choosen_sections).values('id', 'number'))
            response = {'appartments':appartments}
            return JsonResponse(response, safe=False)


        if self.request.is_ajax() and self.request.method == 'GET' and self.request.GET.get('ajax_indicator') == 'get_personal_account_per_appartment':
            if self.request.GET.get('appartment') != "empty_appartment":
                appartment_id = self.request.GET.get('appartment')
                choosen_personal_account = list(PersonalAccount.objects.filter(appartment_account__id=appartment_id)\
                                                .values('id', 'number', 'appartment_account__owner_user__full_name',\
                                                        'appartment_account__owner_user__id','appartment_account__owner_user__phone'))
                choosen_tariff = list(Tariff.objects.filter(appartment_tariff__id=appartment_id).values('id'))
                response = {'personal_account': choosen_personal_account,
                            'choosen_tariff': choosen_tariff}
                return JsonResponse(response, safe=False)
        

        if self.request.is_ajax() and self.request.method == 'GET' and self.request.GET.get('ajax_indicator') == 'add_counters_readings':
            Q_list = []
            tariff_id = self.request.GET.get('tariff_id')
            appartment_id = self.request.GET.get('appartment_data')
            Q_list.append(Q(tariff__id=tariff_id))
            Q_list.append(Q(tariff__appartment_tariff=appartment_id))
            existed_counters_readings = list(CounterReadings.objects.filter(appartment__id=appartment_id).values('utility_service__id').distinct())
            existed_counters = []
            for counter_marker in  existed_counters_readings: existed_counters.append(counter_marker['utility_service__id'])
            Q_list.append(reduce(operator.or_, (Q(utility_service__id=counter_number) for counter_number in existed_counters)))

            tariff_cell_data = list(TariffCell.objects.filter(*Q_list)\
                                            .values('utility_service__id', 'utility_service__unit_of_measure__id', 'cost_per_unit'))
            response = {'tariff_cell_data': tariff_cell_data}
            return JsonResponse(response, safe=False)


        if self.request.is_ajax() and self.request.method == 'GET' and self.request.GET.get('ajax_indicator') == 'add_all_utilities_using_tariff':
            tariff_id = self.request.GET.get('tariff_id')
            counter_tariff_cell_data = list(TariffCell.objects.filter(tariff__id=tariff_id).values('utility_service__id', 'utility_service__unit_of_measure__id', 'cost_per_unit'))
            response = {'counter_tariff_cell_data': counter_tariff_cell_data}
            return JsonResponse(response, safe=False)



        else:
            return super().get(request, *args, **kwargs)


    def post(self, request, *args, **Kwargs):
        main_form = AddReceiptForm(request.POST, prefix='main_form')        
        receipt_cell_formset = ReceiptCellFormset(request.POST, prefix='receipt_cell_formset')
        if main_form.is_valid() and receipt_cell_formset.is_valid():
            return self.form_valid(main_form, receipt_cell_formset)
        else:

            if main_form.errors:
                for field, error in main_form.errors.items():
                    print(f'{field}: {error}')

            if receipt_cell_formset.errors:
                for receipt_form in receipt_cell_formset:
                    for field, error in receipt_form.errors.items():
                        print(f'{field}: {error}')

        
    def form_valid(self, main_form, receipt_cell_formset):
        mainform_instance = main_form.save()
        for receipt_cell_form in receipt_cell_formset:
            receipt_cell = receipt_cell_form.save(commit=False)
            receipt_cell.receipt = mainform_instance
            receipt_cell.save()
        success_url = self.success_url
        messages.success(self.request, f"Квитанция создана!")
        return HttpResponseRedirect(success_url)
    

# DETAIL VIEW----------------------------------------------------------------------------------------------------
class ReceiptCardView(DetailView):
    queryset = Receipt.objects.all()
    template_name = "receipts/receipt_card.html"
    context_object_name = 'receipt'
    


class ReceiptTemplateListView(FormView):
    form_class = ReceiptTemplateListForm
    template_name = "receipts/receipt_template_list.html"
    success_url = reverse_lazy('receipts:receipt_list')

    def form_valid(self, form):
        receipt_id = self.kwargs['pk']
        template_id = form.cleaned_data['templates_list']

        if 'print_xls_doc' in self.request.POST:
            response = return_xlm_receipt(receipt_id, template_id)
            return response

        elif 'send_to_email_pdf' in self.request.POST:
            print('You send email!')
            response = return_pdf_receipt(receipt_id, template_id)
            # response = HttpResponseRedirect(self.success_url)
        
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)     
        receipt_id = self.kwargs['pk']
        context['receipt_id'] = receipt_id
        return context
#########################################################################################################
#########################################################################################################
###############################----------------WORK-----------###########################################
#########################################################################################################
#########################################################################################################

import os
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders



def link_callback(uri, rel):
        """
        Convert HTML URIs to absolute system paths so xhtml2pdf can access those
        resources
        """
        result = finders.find(uri)
        if result:
            if not isinstance(result, (list, tuple)):
                    result = [result]
            result = list(os.path.realpath(path) for path in result)
            path=result[0]
        else:
            sUrl = settings.STATIC_URL        # Typically /static/
            sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
            mUrl = settings.MEDIA_URL         # Typically /media/
            mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

            if uri.startswith(mUrl):
                    path = os.path.join(mRoot, uri.replace(mUrl, ""))
            elif uri.startswith(sUrl):
                    path = os.path.join(sRoot, uri.replace(sUrl, ""))
            else:
                    return uri

        # make sure that file exists
        if not os.path.isfile(path):
                raise Exception(
                        'media URI must start with %s or %s' % (sUrl, mUrl)
                )
        return path


def return_pdf_receipt(receipt_id, template_id):
    print(f'Receipt id: {receipt_id}')
    print(f'Template id: {template_id}')

    receipt_data = list(Receipt.objects\
                            .filter(id=receipt_id)\
                            .values('id', 'appartment__personal_account__number',\
                                    'number', 'payment_due'))
    

    template_path = 'receipts/test_templates.html'
    context = {'receipt_id': receipt_id, 'template_id': template_id}

    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response, link_callback=link_callback)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

    # with NamedTemporaryFile() as tmp:
        # template_for_pdf = 'receipts/test_templates.html'
        # context = {'receipt_id': receipt_id, 'template_id': template_id}

        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        
    #     template = get_template(template_for_pdf)
    #     html = template.render(context)
    #     # new:
    #     result = BytesIO()
    #     pisa_status = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result, link_callback=link_callback)


    #     # pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
        
    #     response = HttpResponse(result, content_type='application/pdf')
    #     response['Content-Disposition'] = 'attachment; filename="report.pdf"'

    #     # if pisa_status.err:
    #         # return HttpResponse('We had some errors <pre>' + html + '</pre>')
    #         # return HttpResponse(result.getvalue(), content_type='application/pdf')
        
    #     # return None
    #     return response

    # #     source_html = render_to_string(template_for_pdf, context)
    #     pisa.CreatePDF(source_html, dest=tmp, encoding='utf-8', link_callback=link_callback)
    #     tmp.seek(0)
    #     stream = tmp.read()
    #     response = HttpResponse(stream, content_type='application/pdf')
    #     response['Content-Disposition'] = f'attachment; testfile.pdf'

    # return response



    # good work but without css and bootstrap
    # template_path = 'receipts/test_templates.html'
    # context = {'receipt_id': receipt_id, 'template_id': template_id}

    # template = get_template(template_path)
    # html = template.render(context)



    # pisa_status = pisa.CreatePDF(
    #     html, desc=response,
    #     link_callback=link_callback 
    # )

    # response = HttpResponse(html, content_type='application/pdf')
    # response['Content-Disposition'] = f'attachment; testfile.pdf'

    # if pisa_status.err:
    #    return HttpResponse('We had some errors <pre>' + html + '</pre>')

    # return response




#########################################################################################################
#########################################################################################################
###############################----------------END#WORK-------###########################################
#########################################################################################################
#########################################################################################################


def return_xlm_receipt(receipt_id, template_id):
    receipt = Receipt.objects.get(id=receipt_id)
    template = ReceiptTemplate.objects.get(id=template_id)

    formated_month = format_date(receipt.from_date, 'LLLL Y', locale='ru')
    receipt_data_dictionary = {'%' + 'accountNumber' + '%': receipt.appartment.personal_account.number,
                               '%' + 'payCompany' + '%': 'TEMPORARY EMPTY DATA. CHANGE AFTER CREATING REQUISITE',
                               '%' + 'invoiceNumber' + '%': receipt.number,
                               '%' + 'invoiceDate' + '%': receipt.payment_due.strftime("%d.%m.%Y"),
                               '%' + 'invoiceAddress' + '%': f'{receipt.appartment.owner_user.full_name}, {receipt.appartment.house.address}, {receipt.appartment.number} квартира',
                               '%' + 'total' + '%': receipt.total_sum,
                               '%' + 'accountBalance' + '%': f'{receipt.appartment.personal_account.balance}',
                               '%' + 'totalDebt' + '%': f'{-(receipt.appartment.personal_account.balance - receipt.total_sum)}',
                               '%' + 'invoiceMonth' + '%': f'{formated_month}',
                               '%' + 'serviceTotal' + '%': f'{receipt.total_sum}',
                               }
                                
    # my workbook
    current_template = load_workbook(filename=str(template.receipt_template.file))
    
    # my work sheet
    current_sheet = current_template.active

    # ----------------------------------------------------------------------------------------
    # get all receipt sells
    receipt_cells = list(ReceiptCell.objects.filter(receipt=receipt)\
                                        .values('utility_service__title',\
                                                 'cost_per_unit',
                                                 'unit_of_measure__title',
                                                 'consumption',
                                                 'cost'))
    
    current_row = 19
    number_of_receipts = len(receipt_cells)
    maximum_rows = current_row + number_of_receipts + 5

    # work with cells of receipt
    for row in current_sheet.iter_rows(min_row=1, min_col=1, max_row=maximum_rows, max_col=12):
        for cell in row:
            if cell.value in receipt_data_dictionary.keys():
                current_sheet[cell.coordinate] = receipt_data_dictionary[cell.value]

    

    # styles for deifferent types of cells

    simple_cell_style = NamedStyle('cell_style')
    simple_cell_style.font = Font(size=12, italic=False)
    simple_cell_style.border = Border(
        left=Side(style="thin", color="00333333"),
        right=Side(style="thin", color="00333333"),
        top=Side(style='thick', color="00333333"),
        bottom=Side(style='thick', color="00333333"),
    )
    simple_cell_style.alignment = Alignment(horizontal='center', vertical='center')
    # simple_cell_style.row_dimensions[1].height = 70
    current_template.add_named_style(simple_cell_style)


    for receipt_cell in receipt_cells:

        current_sheet.insert_rows(current_row)
        current_sheet.row_dimensions[current_row].height = 25

        # utility services cell
        current_sheet[f'A{current_row}'] =  receipt_cell['utility_service__title']
        current_sheet[f'A{current_row}'].style = simple_cell_style
        current_sheet.merge_cells(f'A{current_row}:B{current_row}')
        
        # cost per unit cell
        current_sheet[f'C{current_row}'] = receipt_cell['cost_per_unit']
        current_sheet[f'C{current_row}'].style = simple_cell_style
        current_sheet[f'C{current_row}'].number_format = '#,##0.00'
        current_sheet.merge_cells(f'C{current_row}:D{current_row}')

        # unit of measure cell
        current_sheet[f'E{current_row}'] = receipt_cell['unit_of_measure__title']
        current_sheet[f'E{current_row}'].style = simple_cell_style
        current_sheet.merge_cells(f'E{current_row}:F{current_row}')

        # consumption cell
        current_sheet[f'G{current_row}'] = receipt_cell['consumption']
        current_sheet[f'G{current_row}'].style = simple_cell_style
        current_sheet[f'G{current_row}'].number_format = '#,##0.00'
        current_sheet.merge_cells(f'G{current_row}:H{current_row}')

        # cost cell
        current_sheet[f'I{current_row}'] = receipt_cell['cost']
        current_sheet[f'I{current_row}'].style = simple_cell_style
        current_sheet[f'I{current_row}'].number_format = '#,##0.00'
        current_sheet.merge_cells(f'I{current_row}:K{current_row}')


        current_sheet[f'K{current_row}'].border = Border(top=Side(style='thick'), \
                                                        left=Side(style='thick'), \
                                                        right=Side(style='thick'), \
                                                        bottom=Side(style='thick'))

        # counterpdf_receipt(receipt_id, template_id):
#     pass
        current_row += 1

    # total data logic 
    current_sheet.merge_cells(start_row=current_row,\
                                start_column=1,\
                                end_row=current_row+3,\
                                end_column=2)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=3,\
                                end_row=current_row+3,\
                                end_column=4)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=5,\
                                end_row=current_row+3,\
                                end_column=6)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=7,\
                                end_row=current_row+3,\
                                end_column=8)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=9,\
                                end_row=current_row+3,\
                                end_column=11)

    current_sheet[f'A{current_row}'] = ''

    current_sheet[f'G{current_row}'] = 'РАЗОМ'
    current_sheet[f'G{current_row}'].font = Font(size=12, italic=False, bold=True, color="00000000")
    current_sheet[f'G{current_row}'].alignment = Alignment(horizontal='right')

    current_sheet[f'I{current_row}'] = receipt.total_sum
    current_sheet[f'I{current_row}'].font = Font(size=12, italic=False, bold=True, color="00000000")
    current_sheet[f'I{current_row}'].alignment = Alignment(horizontal='right')
    
    with NamedTemporaryFile() as tmp:
        current_template.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        response = HttpResponse(stream, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename=invoice_{receipt.number}_{receipt.payment_due.strftime("%d.%m.%Y")}.xlsx'

    return response

# def return_pdf_receipt(receipt_id, template_id):
#     pass



class ReceiptTemplateEditeView(FormView):
    template_name = "receipts/receipt_template_edite.html"
    templates = ReceiptTemplate.objects.all()
    form_class = ReceiptTeplateEditeForm
    success_url = reverse_lazy('receipts:receipt_template_edite_view')


    def post(self, request, *args, **Kwargs):
        template_edit_formset = ReceiptTeplateEditeFormSet(request.POST,\
                                                           request.FILES,\
                                                            queryset=self.templates,\
                                                            prefix="templates")
        
        if template_edit_formset.is_valid():
            return self.form_valid(template_edit_formset)
        else:
            if template_edit_formset.errors:
                for receipt_form in template_edit_formset:
                    for field, error in receipt_form.errors.items():
                        print(f'{field}: {error}')



    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)     
        context['receipt_edit_formset'] = ReceiptTeplateEditeFormSet(queryset=self.templates,\
                                                                    prefix="templates")
        return context
    

    def form_valid(self, template_edit_formset):
        template_edit_formset.save()
        success_url = self.success_url
        messages.success(self.request, f"Изменения в шаблоны внесены!")
        return HttpResponseRedirect(success_url)