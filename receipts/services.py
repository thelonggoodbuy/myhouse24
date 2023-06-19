import os
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib.staticfiles import finders
from babel.dates import format_date
from openpyxl import load_workbook
from statements.models import Statement
from django.core.files import File

from openpyxl.styles import Side, Border, Font, Alignment, NamedStyle
from openpyxl import Workbook
from tempfile import NamedTemporaryFile

from .models import Receipt, ReceiptTemplate, ReceiptCell, Requisite
import pdfkit

from django.template import Context, Template
from .tasks import send_pdf_receipt
import base64

from django.core.files.storage import FileSystemStorage

from django.templatetags.static import static

from django.core.mail import send_mail, EmailMessage

from appartments.models import PersonalAccount





def return_pdf_receipt(receipt_id, template_id):

    # --------------------------------------------------------------------
    # START context data logic
    # --------------------------------------------------------------------
    receipt = Receipt.objects.get(id=receipt_id)
    template = ReceiptTemplate.objects.get(id=template_id)

    formated_month = format_date(receipt.from_date, 'LLLL Y', locale='ru')

    requiste = Requisite.objects.first().company_title
    context = {'account_number': receipt.appartment.personal_account.number, 
                'pay_company': requiste,
                'invoice_number': receipt.number, 
                'invoice_date': receipt.payment_due.strftime("%d.%m.%Y"),
                'invoice_address': f'{receipt.appartment.owner_user.full_name}, {receipt.appartment.house.address}, {receipt.appartment.number}',
                'total': receipt.total_sum,
                'account_balance': receipt.appartment.personal_account.balance,
                'total_debt': f'{-(receipt.appartment.personal_account.balance - receipt.total_sum)}',
                'invoice_month': f'{formated_month}',
                'service_total': f'{receipt.total_sum}',
            }
    
    context['receipt_cells'] = list(ReceiptCell.objects.filter(receipt=receipt)\
                                        .values('utility_service__title',\
                                                 'cost_per_unit',
                                                 'unit_of_measure__title',
                                                 'consumption',
                                                 'cost'))
   
    template_path = 'pdf_templates/test_templates.html'

    send_pdf_receipt.delay(receipt.appartment.personal_account.number,
                    receipt.appartment.owner_user.full_name,
                    receipt.appartment.owner_user.email,
                    template_path, context)

    return None

def return_xlm_receipt(receipt_id, template_id):
    receipt = Receipt.objects.get(id=receipt_id)
    template = ReceiptTemplate.objects.get(id=template_id)

    requiste = Requisite.objects.first().company_title
    formated_month = format_date(receipt.from_date, 'LLLL Y', locale='ru')
    receipt_data_dictionary = {'%' + 'accountNumber' + '%': receipt.appartment.personal_account.number,
                               '%' + 'payCompany' + '%': requiste,
                               '%' + 'invoiceNumber' + '%': receipt.number,
                               '%' + 'invoiceDate' + '%': receipt.payment_due.strftime("%d.%m.%Y"),
                               '%' + 'invoiceAddress' + '%': f'{receipt.appartment.owner_user.full_name}, {receipt.appartment.house.address}, {receipt.appartment.number} квартира',
                               '%' + 'total' + '%': receipt.total_sum,
                               '%' + 'accountBalance' + '%': f'{receipt.appartment.personal_account.balance}',
                               '%' + 'totalDebt' + '%': f'{-(receipt.appartment.personal_account.balance - receipt.total_sum)}',
                               '%' + 'invoiceMonth' + '%': f'{formated_month}',
                               '%' + 'serviceTotal' + '%': f'{receipt.total_sum}',
                               }
                                
    # my workbook
    current_template = load_workbook(filename=str(template.receipt_template.file))
    
    # my work sheet
    current_sheet = current_template.active

    # ----------------------------------------------------------------------------------------
    # get all receipt sells
    receipt_cells = list(ReceiptCell.objects.filter(receipt=receipt)\
                                        .values('utility_service__title',\
                                                 'cost_per_unit',
                                                 'unit_of_measure__title',
                                                 'consumption',
                                                 'cost'))
    
    current_row = 19
    number_of_receipts = len(receipt_cells)
    maximum_rows = current_row + number_of_receipts + 5

    # work with cells of receipt
    for row in current_sheet.iter_rows(min_row=1, min_col=1, max_row=maximum_rows, max_col=12):
        for cell in row:
            if cell.value in receipt_data_dictionary.keys():
                current_sheet[cell.coordinate] = receipt_data_dictionary[cell.value]

    

    # styles for deifferent types of cells
    simple_cell_style = NamedStyle('cell_style')
    simple_cell_style.font = Font(size=12, italic=False)
    simple_cell_style.border = Border(
        left=Side(style="thin", color="00333333"),
        right=Side(style="thin", color="00333333"),
        top=Side(style='thick', color="00333333"),
        bottom=Side(style='thick', color="00333333"),
    )
    simple_cell_style.alignment = Alignment(horizontal='center', vertical='center')
    current_template.add_named_style(simple_cell_style)


    for receipt_cell in receipt_cells:

        current_sheet.insert_rows(current_row)
        current_sheet.row_dimensions[current_row].height = 25

        # utility services cell
        current_sheet[f'A{current_row}'] =  receipt_cell['utility_service__title']
        current_sheet[f'A{current_row}'].style = simple_cell_style
        current_sheet.merge_cells(f'A{current_row}:B{current_row}')
        
        # cost per unit cell
        current_sheet[f'C{current_row}'] = receipt_cell['cost_per_unit']
        current_sheet[f'C{current_row}'].style = simple_cell_style
        current_sheet[f'C{current_row}'].number_format = '#,##0.00'
        current_sheet.merge_cells(f'C{current_row}:D{current_row}')

        # unit of measure cell
        current_sheet[f'E{current_row}'] = receipt_cell['unit_of_measure__title']
        current_sheet[f'E{current_row}'].style = simple_cell_style
        current_sheet.merge_cells(f'E{current_row}:F{current_row}')

        # consumption cell
        current_sheet[f'G{current_row}'] = receipt_cell['consumption']
        current_sheet[f'G{current_row}'].style = simple_cell_style
        current_sheet[f'G{current_row}'].number_format = '#,##0.00'
        current_sheet.merge_cells(f'G{current_row}:H{current_row}')

        # cost cell
        current_sheet[f'I{current_row}'] = receipt_cell['cost']
        current_sheet[f'I{current_row}'].style = simple_cell_style
        current_sheet[f'I{current_row}'].number_format = '#,##0.00'
        current_sheet.merge_cells(f'I{current_row}:K{current_row}')


        current_sheet[f'K{current_row}'].border = Border(top=Side(style='thick'), \
                                                        left=Side(style='thick'), \
                                                        right=Side(style='thick'), \
                                                        bottom=Side(style='thick'))

        current_row += 1


    current_sheet.merge_cells(start_row=current_row,\
                                start_column=1,\
                                end_row=current_row+3,\
                                end_column=2)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=3,\
                                end_row=current_row+3,\
                                end_column=4)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=5,\
                                end_row=current_row+3,\
                                end_column=6)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=7,\
                                end_row=current_row+3,\
                                end_column=8)

    current_sheet.merge_cells(start_row=current_row,\
                                start_column=9,\
                                end_row=current_row+3,\
                                end_column=11)

    current_sheet[f'A{current_row}'] = ''

    current_sheet[f'G{current_row}'] = 'РАЗОМ'
    current_sheet[f'G{current_row}'].font = Font(size=12, italic=False, bold=True, color="00000000")
    current_sheet[f'G{current_row}'].alignment = Alignment(horizontal='right')

    current_sheet[f'I{current_row}'] = receipt.total_sum
    current_sheet[f'I{current_row}'].font = Font(size=12, italic=False, bold=True, color="00000000")
    current_sheet[f'I{current_row}'].alignment = Alignment(horizontal='right')
    
    with NamedTemporaryFile() as tmp:
        current_template.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        response = HttpResponse(stream, content_type='application/vnd.ms-excel')
        print(response)
        response['Content-Disposition'] = f'attachment; filename=invoice_{receipt.number}_{receipt.payment_due.strftime("%d.%m.%Y")}.xlsx'

    return response



def return_xlm_list_of_statements():
    
    # my workbook
    current_template = Workbook()
    current_sheet = current_template.active

    statements = list(Statement.objects.all().values('number', 'date', 'type_of_statement',\
                                                    'type_of_paynent_item__title', 'summ',\
                                                    'personal_account__appartment_account__owner_user__full_name',\
                                                    'personal_account__number'))

    current_row = 1
    current_sheet[f'A{current_row}'] = '#'
    current_sheet[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'B{current_row}'] = 'Дата'
    current_sheet[f'B{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'C{current_row}'] = 'Приход/расход'
    current_sheet[f'C{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'D{current_row}'] = 'Статьи'
    current_sheet[f'D{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'E{current_row}'] = 'Квитанции'
    current_sheet[f'E{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'F{current_row}'] = 'Услуга'
    current_sheet[f'F{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'G{current_row}'] = 'Сумма'
    current_sheet[f'G{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'H{current_row}'] = 'Валюта'
    current_sheet[f'H{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'J{current_row}'] = 'Владелец'
    current_sheet[f'J{current_row}'].alignment = Alignment(horizontal='center')
    current_sheet[f'K{current_row}'] = 'Лицевой счет'
    current_sheet[f'K{current_row}'].alignment = Alignment(horizontal='center')


    current_sheet.column_dimensions['A'].width = 60
    current_sheet.column_dimensions['B'].width = 40
    current_sheet.column_dimensions['C'].width = 40
    current_sheet.column_dimensions['D'].width = 40
    current_sheet.column_dimensions['E'].width = 40
    current_sheet.column_dimensions['F'].width = 10
    current_sheet.column_dimensions['G'].width = 10
    current_sheet.column_dimensions['H'].width = 10
    current_sheet.column_dimensions['J'].width = 40
    current_sheet.column_dimensions['K'].width = 40


    current_row = 1 + 1

    for statement_cell in statements:
        # formating statements data
        simple_type_of_statement = statement_cell['type_of_statement']
        if simple_type_of_statement == 'arrival':
            statement_cell['type_of_statement'] = 'Приход'
        else:
            statement_cell['type_of_statement'] = 'Расход'


        current_sheet[f'A{current_row}'] =  statement_cell['number']
        current_sheet[f'A{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'B{current_row}'] =  str(statement_cell['date'])
        current_sheet[f'B{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'C{current_row}'] =  statement_cell['type_of_statement']
        current_sheet[f'C{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'D{current_row}'] =  statement_cell['type_of_paynent_item__title']
        current_sheet[f'D{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'G{current_row}'] =  statement_cell['summ']
        current_sheet[f'G{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'H{current_row}'] = 'UAH'
        current_sheet[f'H{current_row}'].alignment = Alignment(horizontal='center')
        
        current_sheet[f'J{current_row}'] =  statement_cell['personal_account__appartment_account__owner_user__full_name']
        current_sheet[f'J{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'K{current_row}'] =  statement_cell['personal_account__number']
        current_sheet[f'K{current_row}'].alignment = Alignment(horizontal='center')

        current_row += 1

    
    with NamedTemporaryFile() as tmp:
        current_template.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        response = HttpResponse(stream, content_type='application/vnd.ms-excel')
        print(response)
        response['Content-Disposition'] = f'attachment; filename=list_of_statements.xlsx'

    return response



def return_xlm_statement_data(pk):
    current_template = Workbook()
    current_sheet = current_template.active

    statement = Statement.objects.get(id=pk)

    current_sheet.column_dimensions['A'].width = 60
    current_sheet.column_dimensions['B'].width = 60

    current_sheet[f'A1'] = 'Платеж'
    current_sheet[f'A1'].alignment = Alignment(horizontal='center')
    current_sheet[f'B1'] = f'#{ statement.number }'
    current_sheet[f'B1'].alignment = Alignment(horizontal='center')
    
    current_sheet[f'A2'] = 'Дата'
    current_sheet[f'A2'].alignment = Alignment(horizontal='center')
    current_sheet[f'B2'] = statement.date
    current_sheet[f'B2'].alignment = Alignment(horizontal='center')
    
    current_sheet[f'A3'] = 'Владелец квартиры'
    current_sheet[f'A3'].alignment = Alignment(horizontal='center')
    current_sheet[f'B3'] = statement.personal_account.appartment_account.owner_user.full_name
    current_sheet[f'B3'].alignment = Alignment(horizontal='center')

    current_sheet[f'A4'] = 'Лицевой счет'
    current_sheet[f'A4'].alignment = Alignment(horizontal='center')
    current_sheet[f'B4'] = statement.personal_account.number
    current_sheet[f'B4'].alignment = Alignment(horizontal='center')

    current_sheet[f'A5'] = 'Приход/расход'
    current_sheet[f'A5'].alignment = Alignment(horizontal='center')
    if statement.type_of_statement == 'arrival':
        current_sheet[f'B5'] = 'Приход'
    else:
        current_sheet[f'B5'] = 'Расход'
    current_sheet[f'B5'].alignment = Alignment(horizontal='center')

    current_sheet[f'A6'] = 'Статус'
    current_sheet[f'A6'].alignment = Alignment(horizontal='center')
    if statement.checked == 'True':
        current_sheet[f'B6'] = 'Проведен'
    else:
        current_sheet[f'B6'] = 'Не проведен'
    current_sheet[f'B5'].alignment = Alignment(horizontal='center')

    current_sheet[f'A6'] = 'Статья'
    current_sheet[f'A6'].alignment = Alignment(horizontal='center')
    if statement.type_of_paynent_item:
        current_sheet[f'B6'] = statement.type_of_paynent_item.title
    else:
        current_sheet[f'B6'] = 'Не указана'
    current_sheet[f'B6'].alignment = Alignment(horizontal='center')

    current_sheet[f'A7'] = 'Квитанция'
    current_sheet[f'A7'].alignment = Alignment(horizontal='center')

    current_sheet[f'A8'] = 'Услуга'
    current_sheet[f'A8'].alignment = Alignment(horizontal='center')
    
    current_sheet[f'A9'] = 'Сумма'
    current_sheet[f'A9'].alignment = Alignment(horizontal='center')
    current_sheet[f'B9'] = statement.summ
    current_sheet[f'B9'].alignment = Alignment(horizontal='center')

    current_sheet[f'A10'] = 'Валюта'
    current_sheet[f'A10'].alignment = Alignment(horizontal='center')
    current_sheet[f'B10'] = 'UAH'
    current_sheet[f'B10'].alignment = Alignment(horizontal='center')

    current_sheet[f'A11'] = 'Комментарий'
    current_sheet[f'A11'].alignment = Alignment(horizontal='center')
    current_sheet[f'B11'] = statement.comment
    current_sheet[f'B11'].alignment = Alignment(horizontal='center')

    current_sheet[f'A11'] = 'Менеджер'
    current_sheet[f'A11'].alignment = Alignment(horizontal='center')
    if statement.manager:
        current_sheet[f'B11'] = statement.manager.full_name
    else:
        current_sheet[f'B11'] = 'Не указана'
    current_sheet[f'B11'].alignment = Alignment(horizontal='center')



    with NamedTemporaryFile() as tmp:
        current_template.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        response = HttpResponse(stream, content_type='application/vnd.ms-excel')
        print(response)
        response['Content-Disposition'] = f'attachment; filename=statements_{statement.number}.xlsx'

    return response



def return_xlm_statement_data():
    current_template = Workbook()
    current_sheet = current_template.active

    personal_accounts = list(PersonalAccount.objects.select_related('appartment_account')\
                                                    .all()\
                                                    .values('number', 'status',\
                                                            'appartment_account__house__title',\
                                                            'appartment_account__sections__title',\
                                                            'appartment_account__number',\
                                                            'appartment_account__owner_user__full_name',\
                                                            'balance'))

    current_sheet.column_dimensions['A'].width = 60
    current_sheet.column_dimensions['B'].width = 15
    current_sheet.column_dimensions['C'].width = 25
    current_sheet.column_dimensions['D'].width = 15
    current_sheet.column_dimensions['E'].width = 15
    current_sheet.column_dimensions['F'].width = 60
    current_sheet.column_dimensions['G'].width = 15

    current_sheet['A1'] = 'Лицевой счет'
    current_sheet['A1'].alignment = Alignment(horizontal='center')
    current_sheet['B1'] = 'Статус'
    current_sheet['B1'].alignment = Alignment(horizontal='center')
    current_sheet['C1'] = 'Дом'
    current_sheet['C1'].alignment = Alignment(horizontal='center')
    current_sheet['D1'] = 'Секция'
    current_sheet['D1'].alignment = Alignment(horizontal='center')
    current_sheet['E1'] = 'Квартира'
    current_sheet['E1'].alignment = Alignment(horizontal='center')
    current_sheet['F1'] = 'Владелец'
    current_sheet['F1'].alignment = Alignment(horizontal='center')
    current_sheet['G1'] = 'Остаток'
    current_sheet['G1'].alignment = Alignment(horizontal='center')

    current_row = 1 + 1

    for account in personal_accounts:
    
        current_sheet[f'A{current_row}'] =  account['number']
        current_sheet[f'A{current_row}'].alignment = Alignment(horizontal='center')

        if account['status'] == 'active':
            current_sheet[f'B{current_row}'] = 'Активный'
        else:
            current_sheet[f'B{current_row}'] = 'Неактивный'
        current_sheet[f'B{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'C{current_row}'] =  account['appartment_account__house__title']
        current_sheet[f'C{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'D{current_row}'] =  account['appartment_account__sections__title']
        current_sheet[f'D{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'E{current_row}'] =  account['appartment_account__number']
        current_sheet[f'E{current_row}'].alignment = Alignment(horizontal='center')

        current_sheet[f'F{current_row}'] = account['appartment_account__owner_user__full_name']
        current_sheet[f'F{current_row}'].alignment = Alignment(horizontal='center')
        
        current_sheet[f'G{current_row}'] =  account['balance']
        current_sheet[f'G{current_row}'].alignment = Alignment(horizontal='center')

        current_row += 1
    
    

    with NamedTemporaryFile() as tmp:
        current_template.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        response = HttpResponse(stream, content_type='application/vnd.ms-excel')
        print(response)
        response['Content-Disposition'] = f'attachment; filename=all_personal_accounts_data.xlsx'

    return response




def cabinet_download_pdf_receipt(receipt_id):

    receipt = Receipt.objects.get(id=receipt_id)

    print(receipt.number)

    formated_month = format_date(receipt.from_date, 'LLLL Y', locale='ru')


    requiste = Requisite.objects.first().company_title

    context = {}

    if receipt.appartment.personal_account: context['account_number'] = receipt.appartment.personal_account.number
    if requiste: context['pay_company'] = requiste
    if receipt.appartment.personal_account:
        context['account_balance'] = receipt.appartment.personal_account.balance
        context['total_debt'] = f'{-(receipt.appartment.personal_account.balance - receipt.total_sum)}'
    context['invoice_number'] = receipt.number
    context['invoice_date'] = receipt.payment_due.strftime("%d.%m.%Y")
    context['invoice_address'] = f'{receipt.appartment.owner_user.full_name}, {receipt.appartment.house.address}, {receipt.appartment.number}'
    context['total'] = receipt.total_sum
    context['invoice_month'] = f'{formated_month}'
    context['service_total'] = f'{receipt.total_sum}'


    context['receipt_cells'] = list(ReceiptCell.objects.filter(receipt=receipt)\
                                        .values('utility_service__title',\
                                                 'cost_per_unit',
                                                 'unit_of_measure__title',
                                                 'consumption',
                                                 'cost'))
   
    print(context)

    template_path = 'pdf_templates/test_templates.html'

    prerendered_template = get_template(template_path)
    html = prerendered_template.render(context)
    myPdf = pdfkit.from_string(html, False)

    response = HttpResponse(myPdf, content_type='content_type=application/pdf')

    response['Content-Disposition'] = f'attachment; filename=receipt_{receipt.number}.pdf'
    

    return response


import subprocess

def cabinet_download_pdf_receipt_for_printing(receipt_id):

    receipt = Receipt.objects.get(id=receipt_id)

    print(receipt.number)

    formated_month = format_date(receipt.from_date, 'LLLL Y', locale='ru')


    requiste = Requisite.objects.first().company_title

    context = {}

    if receipt.appartment.personal_account: context['account_number'] = receipt.appartment.personal_account.number
    if requiste: context['pay_company'] = requiste
    if receipt.appartment.personal_account:
        context['account_balance'] = receipt.appartment.personal_account.balance
        context['total_debt'] = f'{-(receipt.appartment.personal_account.balance - receipt.total_sum)}'
    context['invoice_number'] = receipt.number
    context['invoice_date'] = receipt.payment_due.strftime("%d.%m.%Y")
    context['invoice_address'] = f'{receipt.appartment.owner_user.full_name}, {receipt.appartment.house.address}, {receipt.appartment.number}'
    context['total'] = receipt.total_sum
    context['invoice_month'] = f'{formated_month}'
    context['service_total'] = f'{receipt.total_sum}'


    context['receipt_cells'] = list(ReceiptCell.objects.filter(receipt=receipt)\
                                        .values('utility_service__title',\
                                                 'cost_per_unit',
                                                 'unit_of_measure__title',
                                                 'consumption',
                                                 'cost'))
   
    print(context)

    template_path = 'pdf_templates/test_templates.html'

    prerendered_template = get_template(template_path)
    html = prerendered_template.render(context)
    myPdf = pdfkit.from_string(html, False)

    response = HttpResponse(myPdf, content_type='content_type=application/pdf')
    
    response['Content-Disposition'] = f'print; filename=receipt_{receipt.number}.pdf'
    # os.startfile(myPdf, "print")
    # lpr =  subprocess.Popen(myPdf, stdin=subprocess.PIPE)
    # lpr.stdin.write(myPdf)

    # with open(myPdf, 'rb') as pdf:
    #     response = HttpResponse(myPdf.read(), content_type='application/pdf')
    #     response['Content-Disposition'] = 'inline;filename=mypdf.pdf'
    

    return response